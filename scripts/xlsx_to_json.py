"""
Converte a planilha de projetos SECOCI (XLSX) para JSON estático.
Trata a exceção do 8º Ano (header na linha 2, colunas diferentes).
"""
import json
import re
import openpyxl
from pathlib import Path

XLSX_PATH = Path(__file__).parent.parent / "PROJETOS SECOCI - GERAL (1).xlsx"
OUTPUT_PATH = Path(__file__).parent.parent / "data" / "projetos.json"

# Mapeamento de nomes de abas para nomes legíveis
SERIES_MAP = {
    "1º ANO": "1º Ano",
    "2º ANO": "2º Ano",
    "3º ANO": "3º Ano",
    "4º ANO": "4º Ano",
    "5º ANO": "5º Ano",
    "6º ANO": "6º Ano",
    "7º ANO": "7º Ano",
    "8º ANO": "8º Ano",
    "9º ANO": "9º Ano",
    "1ª SÉRIE": "1ª Série",
    "2ª SÉRIE": "2ª Série",
}


def normalize_sheet_name(name: str) -> str:
    """Normaliza o nome da aba removendo espaços extras."""
    cleaned = name.strip()
    # Tenta match direto ou case-insensitive
    for key, value in SERIES_MAP.items():
        if cleaned.upper() == key.upper():
            return value
    return cleaned


def split_components(raw: str) -> list[str]:
    """Separa nomes de componentes (alunos) por vírgula, \\n, ou ponto-e-vírgula."""
    if not raw:
        return []
    text = str(raw)
    # Normaliza sequências de escape \\n para newlines reais
    text = text.replace('\\n', '\n')
    # Separa por vírgula, ponto-e-vírgula ou newline
    names = re.split(r'[,;\n]+', text)
    return [n.strip() for n in names if n.strip()]


def parse_standard_sheet(ws, serie_name: str) -> list[dict]:
    """Parseia abas com estrutura padrão: header na linha 1."""
    projects = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        num, titulo, turma_raw, componentes_raw, orientador, *_ = (
            list(row) + [None] * 5
        )[:11]

        # Pula linhas sem título
        if not titulo or not str(titulo).strip():
            continue

        turma = str(turma_raw).strip() if turma_raw else ""
        # Extrair letra da turma: "1º ANO - A" → "A", "1A" → "A", "8B" → "B"
        turma_match = re.search(r'[- ]+([A-Za-z])$', turma)
        if turma_match:
            turma_letter = turma_match.group(1).upper()
        else:
            # Fallback: última letra maiúscula
            turma_letter = re.sub(r'[^A-Za-z]', '', turma)[-1:].upper() if re.sub(r'[^A-Za-z]', '', turma) else ""

        projects.append({
            "id": int(num) if num else None,
            "titulo": str(titulo).strip(),
            "serie": serie_name,
            "turma": turma_letter,
            "componentes": split_components(componentes_raw),
            "orientador": str(orientador).strip() if orientador else "",
        })

    return projects


def parse_8ano_sheet(ws, serie_name: str) -> list[dict]:
    """
    Parseia a aba do 8º Ano que tem estrutura diferente:
    - Header na linha 2
    - Colunas: Nº do projeto | Título | Série/turma | [Área] | Componentes | Orientador | [Link]
    """
    projects = []
    # Detecta onde começa o header
    header_row = 1
    for row_idx in range(1, min(ws.max_row + 1, 5)):
        cell_val = ws.cell(row=row_idx, column=1).value
        if cell_val and "projeto" in str(cell_val).lower():
            header_row = row_idx
            break
        # Se a primeira célula é None e a próxima linha tem "Nº do projeto"
        if cell_val is None:
            next_val = ws.cell(row=row_idx + 1, column=1).value
            if next_val and "projeto" in str(next_val).lower():
                header_row = row_idx + 1
                break

    # Mapeia colunas pelo nome do header
    headers = {}
    for col_idx in range(1, ws.max_column + 1):
        val = ws.cell(row=header_row, column=col_idx).value
        if val:
            headers[str(val).strip().lower()] = col_idx

    # Detecta colunas relevantes
    col_num = headers.get("nº do projeto", headers.get("n° do projeto", 1))
    col_titulo = headers.get("título", headers.get("titulo", 2))
    col_turma = headers.get("série/turma", headers.get("serie/turma", 3))

    # Componentes pode estar em diferentes colunas
    col_componentes = None
    col_orientador = None
    for key, idx in headers.items():
        if "componente" in key or "aluno" in key:
            col_componentes = idx
        if "orientador" in key:
            col_orientador = idx

    # Se não achou "componentes", tenta a coluna 5 (que na aba 8º tem os nomes)
    if col_componentes is None:
        col_componentes = 5
    if col_orientador is None:
        col_orientador = 6

    for row_idx in range(header_row + 1, ws.max_row + 1):
        num = ws.cell(row=row_idx, column=col_num).value
        titulo = ws.cell(row=row_idx, column=col_titulo).value
        turma_raw = ws.cell(row=row_idx, column=col_turma).value
        componentes_raw = ws.cell(row=row_idx, column=col_componentes).value
        orientador = ws.cell(row=row_idx, column=col_orientador).value

        if not titulo or not str(titulo).strip():
            continue

        turma = str(turma_raw).strip() if turma_raw else ""
        turma_letter = re.sub(r'[\dº°ª\s]', '', turma).strip().upper() or ""

        projects.append({
            "id": int(num) if num else None,
            "titulo": str(titulo).strip(),
            "serie": serie_name,
            "turma": turma_letter,
            "componentes": split_components(componentes_raw),
            "orientador": str(orientador).strip() if orientador else "",
        })

    return projects


def main():
    wb = openpyxl.load_workbook(XLSX_PATH, data_only=True)
    all_projects = []

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        serie_name = normalize_sheet_name(sheet_name)
        print(f"Processando: {sheet_name} -> {serie_name}")

        if "8" in sheet_name:
            projects = parse_8ano_sheet(ws, serie_name)
        else:
            projects = parse_standard_sheet(ws, serie_name)

        print(f"  → {len(projects)} projetos encontrados")
        all_projects.extend(projects)

    # Atribui IDs sequenciais para projetos sem ID
    max_id = max((p["id"] for p in all_projects if p["id"]), default=0)
    for p in all_projects:
        if p["id"] is None:
            max_id += 1
            p["id"] = max_id

    # Ordena por série e depois por ID
    serie_order = [
        "1º Ano", "2º Ano", "3º Ano", "4º Ano", "5º Ano",
        "6º Ano", "7º Ano", "8º Ano", "9º Ano",
        "1ª Série", "2ª Série",
    ]
    all_projects.sort(
        key=lambda p: (
            serie_order.index(p["serie"]) if p["serie"] in serie_order else 99,
            p["id"],
        )
    )

    # Salva JSON
    OUTPUT_PATH.parent.mkdir(parents=True, exist_ok=True)
    with open(OUTPUT_PATH, "w", encoding="utf-8") as f:
        json.dump(all_projects, f, ensure_ascii=False, indent=2)

    print(f"\n✅ {len(all_projects)} projetos salvos em {OUTPUT_PATH}")

    # Estatísticas
    series_count = {}
    for p in all_projects:
        series_count[p["serie"]] = series_count.get(p["serie"], 0) + 1
    print("\nResumo por série:")
    for serie, count in series_count.items():
        print(f"  {serie}: {count} projetos")


if __name__ == "__main__":
    main()
